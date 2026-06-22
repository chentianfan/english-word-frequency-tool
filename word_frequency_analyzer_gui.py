#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==============================================================================
# 文件名称：word_frequency_analyzer_pro.py
# 版本：Pro 优化版
# 功能描述：英语单词词频统计与中文释义导出工具（增强版）
# ==============================================================================

import re
import sys
import os
import time
import random
import json
import csv
import pickle
import threading
from collections import Counter
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# 第三方库导入
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("错误：缺少 openpyxl 库，请运行：pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    requests = None

# GUI 支持（可选）
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False


# ==============================================================================
# 多语言支持
# ==============================================================================
class LanguageManager:
    """多语言管理类"""
    
    def __init__(self, language='zh'):
        self.language = language
        self.translations = self._get_translations()
    
    def _get_translations(self):
        trans = {}
        
        # 中文
        trans['zh'] = {
            'app_title': '英语单词词频统计工具 Pro',
            'file_settings': '文件设置',
            'batch_mode': '批量处理模式（处理整个文件夹）',
            'output_mode': '输出模式：',
            'separate_output': '逐个输出',
            'merge_output': '合并输出',
            'input_file': '输入文件：',
            'input_folder': '输入文件夹：',
            'output_file': '输出文件：',
            'output_folder': '输出文件夹：',
            'dict_file': '词典文件：',
            'browse': '浏览...',
            'analysis_options': '分析选项',
            'ignore_stopwords': '忽略停用词',
            'lemmatize': '词形还原',
            'offline_mode': '纯离线模式（不联网）',
            'enable_translation': '获取中文释义',
            'show_level': '显示难度等级',
            'show_roots': '显示词根词缀',
            'filter_no_trans': '只保留有释义的单词',
            'threads': '线程数：',
            'max_words': '最多翻译：',
            'export_format': '导出格式',
            'excel_format': 'Excel (.xlsx)',
            'csv_format': 'CSV (.csv)',
            'anki_format': 'Anki 生词本 (.txt)',
            'start_analysis': '开始分析',
            'about_btn': '公告',
            'language_btn': '语言',
            'progress': '进度',
            'ready': '准备就绪',
            'preparing': '正在准备...',
            'complete': '完成！',
            'run_log': '运行日志',
            'error': '错误',
            'warning': '提示',
            'info': '信息',
            'select_input_file': '请选择输入文件',
            'select_input_folder': '请选择输入文件夹',
            'file_not_exist': '文件不存在：',
            'folder_not_exist': '文件夹不存在：',
            'batch_mode_folder': '批量模式下请选择文件夹',
            'analysis_complete': '分析完成！',
            'output_file_label': '输出文件：',
            'batch_complete': '批量处理完成！',
            'success_count': '成功处理 ',
            'files': ' 个文件',
            'output_folder_label': '输出文件夹：',
            'no_files_found': '没有找到可处理的文件',
            'about_title': '公告 / 关于',
            'about_content': (
                '公 告\n\n'
                '联系邮箱：1602467045@qq.com\n\n'
                '英语单词词频统计工具 Pro\n'
                '版本：v2.0\n\n'
                '功能特点：\n'
                '词频统计\n'
                '中文释义（离线词典）\n'
                '难度分级\n'
                '词根词缀分析\n'
                '批量处理\n'
                '多种格式导出\n\n'
                '如有问题或建议，欢迎邮件联系！'
            ),
            'select_file_title': '选择要分析的文件',
            'all_supported_files': '所有支持的文件',
            'text_files': '文本文件',
            'pdf_files': 'PDF 文件',
            'word_files': 'Word 文档',
            'all_files': '所有文件',
            'select_folder_title': '选择包含文件的文件夹',
            'save_file_title': '保存文件',
            'select_dict_title': '选择词典文件',
            'csv_dict_files': 'CSV 词典文件',
        }
        
        # 英文
        trans['en'] = {
            'app_title': 'Word Frequency Analyzer Pro',
            'file_settings': 'File Settings',
            'batch_mode': 'Batch Mode (Process entire folder)',
            'output_mode': 'Output Mode:',
            'separate_output': 'Separate Output',
            'merge_output': 'Merge Output',
            'input_file': 'Input File:',
            'input_folder': 'Input Folder:',
            'output_file': 'Output File:',
            'output_folder': 'Output Folder:',
            'dict_file': 'Dictionary:',
            'browse': 'Browse...',
            'analysis_options': 'Analysis Options',
            'ignore_stopwords': 'Ignore Stop Words',
            'lemmatize': 'Lemmatization',
            'offline_mode': 'Offline Mode (No Internet)',
            'enable_translation': 'Get Chinese Translation',
            'show_level': 'Show Difficulty Level',
            'show_roots': 'Show Word Roots',
            'filter_no_trans': 'Only Keep Words with Translation',
            'threads': 'Threads:',
            'max_words': 'Max Words:',
            'export_format': 'Export Format',
            'excel_format': 'Excel (.xlsx)',
            'csv_format': 'CSV (.csv)',
            'anki_format': 'Anki Deck (.txt)',
            'start_analysis': 'Start Analysis',
            'about_btn': 'About',
            'language_btn': 'Language',
            'progress': 'Progress',
            'ready': 'Ready',
            'preparing': 'Preparing...',
            'complete': 'Complete!',
            'run_log': 'Run Log',
            'error': 'Error',
            'warning': 'Warning',
            'info': 'Info',
            'select_input_file': 'Please select an input file',
            'select_input_folder': 'Please select an input folder',
            'file_not_exist': 'File does not exist: ',
            'folder_not_exist': 'Folder does not exist: ',
            'batch_mode_folder': 'Please select a folder in batch mode',
            'analysis_complete': 'Analysis complete!',
            'output_file_label': 'Output file: ',
            'batch_complete': 'Batch processing complete!',
            'success_count': 'Successfully processed ',
            'files': ' files',
            'output_folder_label': 'Output folder: ',
            'no_files_found': 'No files found to process',
            'about_title': 'About',
            'about_content': (
                'About\n\n'
                'Contact: 1602467045@qq.com\n\n'
                'Word Frequency Analyzer Pro\n'
                'Version: v2.0\n\n'
                'Features:\n'
                'Word frequency statistics\n'
                'Chinese translation (offline dictionary)\n'
                'Difficulty levels\n'
                'Word root analysis\n'
                'Batch processing\n'
                'Multiple export formats\n\n'
                'Questions or suggestions? Email us!'
            ),
            'select_file_title': 'Select file to analyze',
            'all_supported_files': 'All supported files',
            'text_files': 'Text files',
            'pdf_files': 'PDF files',
            'word_files': 'Word documents',
            'all_files': 'All files',
            'select_folder_title': 'Select folder with files',
            'save_file_title': 'Save file',
            'select_dict_title': 'Select dictionary file',
            'csv_dict_files': 'CSV dictionary files',
        }
        
        return trans
    
    def set_language(self, language):
        if language in self.translations:
            self.language = language
    
    def get(self, key):
        return self.translations.get(self.language, {}).get(key, key)


# ==============================================================================
# 进度条类
# ==============================================================================
class ProgressBar:
    """简单的进度条显示类"""

    def __init__(self, total, prefix='进度', length=50):
        self.total = total
        self.prefix = prefix
        self.length = length
        self.current = 0
        self.start_time = time.time()

    def update(self, current=None):
        if current is not None:
            self.current = current
        else:
            self.current += 1

        percent = self.current / self.total
        filled = int(self.length * percent)
        bar = '█' * filled + '░' * (self.length - filled)

        elapsed = time.time() - self.start_time
        if self.current > 0:
            eta = elapsed / self.current * (self.total - self.current)
            eta_str = f" 剩余: {self._format_time(eta)}"
        else:
            eta_str = ""

        print(f'\r{self.prefix}: |{bar}| {percent*100:.1f}% ({self.current}/{self.total}){eta_str}', end='', flush=True)

        if self.current >= self.total:
            print()

    def _format_time(self, seconds):
        if seconds < 60:
            return f"{seconds:.0f}秒"
        elif seconds < 3600:
            return f"{seconds/60:.1f}分钟"
        else:
            return f"{seconds/3600:.1f}小时"


# ==============================================================================
# 配置管理类
# ==============================================================================
class ConfigManager:
    """配置保存和加载类"""

    def __init__(self, config_file='word_freq_config.json'):
        self.config_file = config_file
        self.default_config = {
            'ignore_stopwords': False,
            'lemmatize': False,
            'use_online_translation': True,
            'max_workers': 3,
            'min_word_length': 2,
            'keep_single_letter_words': True,
            'enable_progress': True,
            'last_input_file': '',
            'last_output_path': '',
            'export_format': 'excel'
        }
        self.config = self._load_config()

    def _load_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                    merged = self.default_config.copy()
                    merged.update(loaded)
                    return merged
            except Exception as e:
                print(f"警告：配置文件加载失败，使用默认配置 ({e})")
                return self.default_config.copy()
        return self.default_config.copy()

    def save(self):
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"警告：配置保存失败 ({e})")

    def get(self, key, default=None):
        return self.config.get(key, default)

    def set(self, key, value):
        self.config[key] = value


# ==============================================================================
# 主分析类
# ==============================================================================
class WordFrequencyAnalyzer:
    """英语单词词频分析器（Pro优化版）"""

    def __init__(self, dict_path=None):
        # 基础配置
        self.min_word_length = 2
        self.keep_single_letter_words = True

        # 翻译相关
        self.use_online_translation = True
        self.max_workers = 3
        self.offline_dict = None
        self.dict_path = dict_path
        self.dict_cache_path = 'ecdict_cache.pkl'

        # 本地小词典
        self.local_dict = {
            'the': '这；那；定冠词',
            'be': '是；存在；成为',
            'to': '到；向；给；不定式符号',
            'of': '…的；属于；关于',
            'and': '和；与；而且',
            'a': '一；一个；不定冠词',
            'in': '在…里；在…内；用',
            'that': '那个；那；引导从句',
            'have': '有；拥有；使',
            'i': '我',
            'it': '它；这；那',
            'for': '为了；给；对于',
            'not': '不；没有',
            'on': '在…上；关于',
            'with': '和…一起；用；带有',
            'he': '他',
            'as': '作为；像；当…时',
            'you': '你；你们',
            'do': '做；干；助动词',
            'at': '在；向；以',
            'this': '这；这个',
            'but': '但是；然而；除了',
            'his': '他的',
            'by': '由；被；通过',
            'from': '从；来自',
            'they': '他们；她们；它们',
            'we': '我们',
            'say': '说；讲',
            'her': '她的；她',
            'she': '她',
            'or': '或者；还是',
            'an': '一；一个（元音前）',
            'will': '将；愿意；意志',
            'my': '我的',
            'one': '一；一个',
            'all': '全部的；所有的',
            'would': '将；愿意；会',
            'there': '那里；在那里',
            'their': '他们的',
            'what': '什么；多么',
            'so': '所以；如此；那么',
            'up': '向上；起来',
            'out': '出去；在外',
            'if': '如果；是否',
            'about': '关于；大约',
            'who': '谁',
            'get': '得到；获得；变得',
            'which': '哪一个；哪些',
            'go': '去；走；进行',
            'me': '我（宾格）',
            'when': '什么时候；当…时',
            'make': '做；制作；使',
            'can': '能；可以；罐头',
            'like': '喜欢；像',
            'time': '时间；次数；时代',
            'just': '只是；刚刚；正好',
            'them': '他们（宾格）',
            'see': '看见；理解',
            'other': '其他的；另一个',
            'than': '比',
            'then': '然后；那么',
            'now': '现在；此刻',
            'look': '看；看起来',
            'only': '只有；仅仅',
            'come': '来；来到',
            'its': '它的',
            'over': '在…上方；超过；结束',
            'think': '想；认为；思考',
            'also': '也；而且',
            'back': '后面；回来；向后',
            'after': '在…之后',
            'use': '使用；用途',
            'two': '二；两个',
            'how': '怎样；如何',
            'our': '我们的',
            'work': '工作；劳动；作品',
            'first': '第一；首先',
            'well': '好；健康的；井',
            'way': '方法；道路；方式',
            'even': '甚至；即使',
            'new': '新的；新鲜的',
            'want': '想要；需要',
            'because': '因为',
            'any': '任何的；一些',
            'these': '这些',
            'give': '给；给予',
            'day': '天；白天；日子',
            'most': '最；大多数',
            'us': '我们（宾格）',
            'is': '是（be的第三人称单数）',
            'are': '是（be的复数和第二人称）',
            'was': '是（be的过去式单数）',
            'were': '是（be的过去式复数）',
            'been': '是（be的过去分词）',
            'has': '有（have的第三人称单数）',
            'had': '有（have的过去式）',
            'does': '做（do的第三人称单数）',
            'did': '做（do的过去式）',
            'am': '是（be的第一人称单数）',
            'people': '人；人们；人民',
            'thing': '东西；事情；事物',
            'man': '男人；人；人类',
            'woman': '女人；妇女',
            'child': '孩子；儿童',
            'world': '世界；地球；领域',
            'life': '生活；生命；人生',
            'hand': '手；指针；帮助',
            'part': '部分；角色；零件',
            'place': '地方；位置；放置',
            'case': '情况；案例；箱子',
            'week': '周；星期',
            'company': '公司；陪伴；连队',
            'system': '系统；体系；制度',
            'program': '程序；节目；计划',
            'question': '问题；疑问；询问',
            'government': '政府；治理',
            'number': '数字；数量；号码',
            'night': '夜晚；晚上',
            'point': '点；要点；指向',
            'home': '家；家庭；回家',
            'water': '水；浇水；水域',
            'room': '房间；空间；余地',
            'mother': '母亲；妈妈',
            'area': '面积；区域；地区',
            'money': '钱；货币；金钱',
            'story': '故事；小说；楼层',
            'fact': '事实；真相',
            'month': '月；月份',
            'lot': '许多；抽签',
            'right': '正确的；右边；权利',
            'study': '学习；研究；书房',
            'book': '书；书籍；预订',
            'eye': '眼睛；眼光；看',
            'job': '工作；职业；任务',
            'word': '单词；话；消息',
            'business': '商业；生意；事务',
            'issue': '问题；议题；发行',
            'side': '边；侧面；方面',
            'kind': '种类；善良的',
            'head': '头；头部；领导',
            'house': '房子；住宅；家庭',
            'service': '服务；服役；业务',
            'friend': '朋友；友人',
            'father': '父亲；爸爸',
            'power': '力量；权力；电力',
            'hour': '小时；钟头',
            'game': '游戏；比赛；猎物',
            'line': '线；路线；行',
            'end': '结束；末端；终点',
            'member': '成员；会员',
            'family': '家庭；家族；家人',
            'problem': '问题；难题',
            'student': '学生；学者；研究者',
        }

        # 停用词列表
        self.stop_words = {
            'a', 'an', 'the',
            'i', 'me', 'my', 'myself', 'we', 'our', 'ours', 'ourselves',
            'you', 'your', 'yours', 'yourself', 'yourselves',
            'he', 'him', 'his', 'himself', 'she', 'her', 'hers', 'herself',
            'it', 'its', 'itself', 'they', 'them', 'their', 'theirs', 'themselves',
            'what', 'which', 'who', 'whom', 'this', 'that', 'these', 'those',
            'am', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
            'have', 'has', 'had', 'having',
            'do', 'does', 'did', 'doing',
            'will', 'would', 'shall', 'should', 'can', 'could', 'may', 'might', 'must',
            'of', 'in', 'on', 'at', 'by', 'for', 'with', 'about', 'against',
            'between', 'into', 'through', 'during', 'before', 'after', 'above',
            'below', 'to', 'from', 'up', 'down', 'in', 'out', 'over', 'under',
            'again', 'further', 'then', 'once',
            'and', 'but', 'or', 'nor', 'so', 'yet', 'both', 'either', 'neither',
            'each', 'every', 'all', 'any', 'few', 'more', 'most', 'other',
            'some', 'such', 'no', 'not', 'only', 'own', 'same', 'than', 'too', 'very',
            'just', 'also', 'now', 'here', 'there', 'when', 'where', 'why', 'how',
            "i'm", "you're", "he's", "she's", "it's", "we're", "they're",
            "i've", "you've", "we've", "they've",
            "i'll", "you'll", "he'll", "she'll", "we'll", "they'll",
            "i'd", "you'd", "he'd", "she'd", "we'd", "they'd",
            "don't", "doesn't", "didn't", "won't", "wouldn't",
            "can't", "couldn't", "shouldn't", "mustn't",
            "isn't", "aren't", "wasn't", "weren't",
            "haven't", "hasn't", "hadn't",
            "that's", "there's", "what's", "who's", "where's", "when's", "why's", "how's",
        }

        # 不规则动词表
        self.irregular_verbs = {
            'am': 'be', 'is': 'be', 'are': 'be',
            'was': 'be', 'were': 'be', 'been': 'be', 'being': 'be',
            'has': 'have', 'had': 'have', 'having': 'have',
            'does': 'do', 'did': 'do', 'done': 'do', 'doing': 'do',
            'goes': 'go', 'went': 'go', 'gone': 'go', 'going': 'go',
            'said': 'say', 'saying': 'say', 'says': 'say',
            'made': 'make', 'making': 'make', 'makes': 'make',
            'took': 'take', 'taken': 'take', 'taking': 'take', 'takes': 'take',
            'came': 'come', 'coming': 'come', 'comes': 'come',
            'saw': 'see', 'seen': 'see', 'seeing': 'see', 'sees': 'see',
            'knew': 'know', 'known': 'know', 'knowing': 'know', 'knows': 'know',
            'got': 'get', 'gotten': 'get', 'getting': 'get', 'gets': 'get',
            'gave': 'give', 'given': 'give', 'giving': 'give', 'gives': 'give',
            'ran': 'run', 'running': 'run', 'runs': 'run',
            'sat': 'sit', 'sitting': 'sit', 'sits': 'sit',
            'stood': 'stand', 'standing': 'stand', 'stands': 'stand',
            'ate': 'eat', 'eaten': 'eat', 'eating': 'eat', 'eats': 'eat',
            'drank': 'drink', 'drunk': 'drink', 'drinking': 'drink', 'drinks': 'drink',
            'slept': 'sleep', 'sleeping': 'sleep', 'sleeps': 'sleep',
            'spoke': 'speak', 'spoken': 'speak', 'speaking': 'speak', 'speaks': 'speak',
            'wrote': 'write', 'written': 'write', 'writing': 'write', 'writes': 'write',
            'read': 'read', 'reading': 'read', 'reads': 'read',
            'bought': 'buy', 'buying': 'buy', 'buys': 'buy',
            'sold': 'sell', 'selling': 'sell', 'sells': 'sell',
            'brought': 'bring', 'bringing': 'bring', 'brings': 'bring',
            'thought': 'think', 'thinking': 'think', 'thinks': 'think',
            'taught': 'teach', 'teaching': 'teach', 'teaches': 'teach',
            'found': 'find', 'finding': 'find', 'finds': 'find',
            'lost': 'lose', 'losing': 'lose', 'loses': 'lose',
            'felt': 'feel', 'feeling': 'feel', 'feels': 'feel',
            'told': 'tell', 'telling': 'tell', 'tells': 'tell',
            'heard': 'hear', 'hearing': 'hear', 'hears': 'hear',
            'kept': 'keep', 'keeping': 'keep', 'keeps': 'keep',
            'let': 'let', 'letting': 'let', 'lets': 'let',
            'put': 'put', 'putting': 'put', 'puts': 'put',
            'set': 'set', 'setting': 'set', 'sets': 'set',
            'began': 'begin', 'begun': 'begin', 'beginning': 'begin', 'begins': 'begin',
            'stopped': 'stop', 'stopping': 'stop', 'stops': 'stop',
            'sent': 'send', 'sending': 'send', 'sends': 'send',
            'built': 'build', 'building': 'build', 'builds': 'build',
            'stayed': 'stay', 'staying': 'stay', 'stays': 'stay',
            'fell': 'fall', 'fallen': 'fall', 'falling': 'fall', 'falls': 'fall',
            'cut': 'cut', 'cutting': 'cut', 'cuts': 'cut',
            'hit': 'hit', 'hitting': 'hit', 'hits': 'hit',
            'met': 'meet', 'meeting': 'meet', 'meets': 'meet',
            'led': 'lead', 'leading': 'lead', 'leads': 'lead',
            'understood': 'understand', 'understanding': 'understand', 'understands': 'understand',
            'watched': 'watch', 'watching': 'watch', 'watches': 'watch',
            'followed': 'follow', 'following': 'follow', 'follows': 'follow',
            'created': 'create', 'creating': 'create', 'creates': 'create',
            'added': 'add', 'adding': 'add', 'adds': 'add',
            'spent': 'spend', 'spending': 'spend', 'spends': 'spend',
            'grew': 'grow', 'grown': 'grow', 'growing': 'grow', 'grows': 'grow',
            'opened': 'open', 'opening': 'open', 'opens': 'open',
            'walked': 'walk', 'walking': 'walk', 'walks': 'walk',
            'won': 'win', 'winning': 'win', 'wins': 'win',
            'offered': 'offer', 'offering': 'offer', 'offers': 'offer',
            'remembered': 'remember', 'remembering': 'remember', 'remembers': 'remember',
            'considered': 'consider', 'considering': 'consider', 'considers': 'consider',
            'appeared': 'appear', 'appearing': 'appear', 'appears': 'appear',
            'waited': 'wait', 'waiting': 'wait', 'waits': 'wait',
            'served': 'serve', 'serving': 'serve', 'serves': 'serve',
            'died': 'die', 'dying': 'die', 'dies': 'die',
            'expected': 'expect', 'expecting': 'expect', 'expects': 'expect',
        }

        # 不规则形容词/副词
        self.irregular_adjectives = {
            'better': 'good', 'best': 'good',
            'worse': 'bad', 'worst': 'bad',
            'more': 'much', 'most': 'much',
            'less': 'little', 'least': 'little',
            'farther': 'far', 'farthest': 'far',
            'further': 'far', 'furthest': 'far',
            'older': 'old', 'oldest': 'old',
            'elder': 'old', 'eldest': 'old',
        }

        # 不规则名词复数
        self.irregular_plurals = {
            'children': 'child',
            'men': 'man',
            'women': 'woman',
            'people': 'person',
            'feet': 'foot',
            'teeth': 'tooth',
            'geese': 'goose',
            'mice': 'mouse',
            'lice': 'louse',
            'oxen': 'ox',
            'phenomena': 'phenomenon',
            'criteria': 'criterion',
            'data': 'datum',
            'bacteria': 'bacterium',
            'fungi': 'fungus',
            'nuclei': 'nucleus',
            'leaves': 'leaf',
            'knives': 'knife',
            'wives': 'wife',
            'lives': 'life',
            'halves': 'half',
            'shelves': 'shelf',
            'wolves': 'wolf',
            'thieves': 'thief',
            'sheep': 'sheep',
            'fish': 'fish',
            'deer': 'deer',
            'series': 'series',
            'species': 'species',
        }

        # 词根词缀数据
        self.word_roots = None

        # 难度等级标签映射
        self.level_labels = {
            'zk': '中考',
            'gk': '高考',
            'cet4': '四级',
            'cet6': '六级',
            'ky': '考研',
            'ielts': '雅思',
            'toefl': '托福',
            'gre': 'GRE',
            'sat': 'SAT',
        }

        # 进度保存相关
        self.enable_progress = True
        self.progress_save_interval = 5
        self._progress_lock = threading.Lock()

        # 请求头设置
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

    # ==========================================================================
    # 单词提取
    # ==========================================================================
    def extract_words(self, text):
        pattern = r"[a-zA-Z]+(?:['-][a-zA-Z]+)*"
        words = re.findall(pattern, text)
        words = [word.lower() for word in words]

        filtered_words = []
        for word in words:
            word_len = len(word)
            if word_len >= self.min_word_length:
                filtered_words.append(word)
            elif word_len == 1 and self.keep_single_letter_words:
                if word in ('a', 'i'):
                    filtered_words.append(word)

        return filtered_words

    # ==========================================================================
    # 词频统计
    # ==========================================================================
    def count_frequency(self, words):
        return Counter(words)

    # ==========================================================================
    # 词形还原
    # ==========================================================================
    def lemmatize_word(self, word):
        word_lower = word.lower()

        if len(word_lower) <= 2:
            return word_lower

        if word_lower in self.irregular_verbs:
            return self.irregular_verbs[word_lower]

        if word_lower in self.irregular_adjectives:
            return self.irregular_adjectives[word_lower]

        if word_lower in self.irregular_plurals:
            return self.irregular_plurals[word_lower]

        # 名词复数
        if word_lower.endswith('ies') and len(word_lower) > 4:
            return word_lower[:-3] + 'y'
        if word_lower.endswith('es') and len(word_lower) > 3:
            if word_lower[-3] in 'sxzo' or word_lower.endswith('ches') or word_lower.endswith('shes'):
                return word_lower[:-2]
        if word_lower.endswith('s') and len(word_lower) > 3:
            if not word_lower.endswith('ss') and not word_lower.endswith('us'):
                return word_lower[:-1]

        # 动词过去式
        if word_lower.endswith('ied') and len(word_lower) > 4:
            return word_lower[:-3] + 'y'
        if word_lower.endswith('ed') and len(word_lower) > 4:
            if word_lower[-3] == word_lower[-4]:
                return word_lower[:-3]
            return word_lower[:-2]

        # 动词现在分词
        if word_lower.endswith('ing') and len(word_lower) > 5:
            if word_lower[-4] == word_lower[-5]:
                return word_lower[:-4]
            if word_lower[-4] != 'e':
                return word_lower[:-3] + 'e'
            return word_lower[:-3]

        # 形容词比较级
        if word_lower.endswith('ier') and len(word_lower) > 4:
            return word_lower[:-3] + 'y'
        if word_lower.endswith('iest') and len(word_lower) > 5:
            return word_lower[:-4] + 'y'
        if word_lower.endswith('er') and len(word_lower) > 4:
            if word_lower[-3] == word_lower[-4]:
                return word_lower[:-3]
            return word_lower[:-2]
        if word_lower.endswith('est') and len(word_lower) > 5:
            if word_lower[-4] == word_lower[-5]:
                return word_lower[:-4]
            return word_lower[:-3]

        # 副词
        if word_lower.endswith('ly') and len(word_lower) > 4:
            return word_lower[:-2]

        return word_lower

    # ==========================================================================
    # 离线词典加载（优化版：带缓存）
    # ==========================================================================
    def _load_offline_dict(self):
        if self.offline_dict is not None:
            return

        print("正在加载离线词典...", end=' ', flush=True)

        dict_file = None
        possible_paths = []

        if self.dict_path:
            possible_paths.append(self.dict_path)

        script_dir = os.path.dirname(os.path.abspath(__file__))
        possible_paths.append(os.path.join(script_dir, 'ecdict.csv'))
        possible_paths.append('ecdict.csv')

        for path in possible_paths:
            if os.path.exists(path):
                dict_file = path
                break

        if dict_file is None:
            print("（未找到词典文件，将仅使用本地小词典）")
            self.offline_dict = {}
            return

        # 检查缓存（缓存文件名基于词典文件名，避免不同词典共用缓存）
        dict_basename = os.path.basename(dict_file)
        cache_file = os.path.join(os.path.dirname(dict_file), f"{dict_basename}_cache.pkl")
        if os.path.exists(cache_file):
            if os.path.getmtime(cache_file) > os.path.getmtime(dict_file):
                try:
                    with open(cache_file, 'rb') as f:
                        self.offline_dict = pickle.load(f)
                    print(f"完成（从缓存加载，共 {len(self.offline_dict)} 个词条）")
                    return
                except Exception:
                    pass

        # 从CSV加载
        self.offline_dict = {}
        try:
            with open(dict_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    word = row.get('word', '').lower()
                    if word:
                        self.offline_dict[word] = {
                            'phonetic': row.get('phonetic', ''),
                            'definition': row.get('definition', ''),
                            'translation': row.get('translation', ''),
                            'pos': row.get('pos', ''),
                            'collins': row.get('collins', ''),
                            'oxford': row.get('oxford', ''),
                            'tag': row.get('tag', ''),
                            'bnc': row.get('bnc', ''),
                            'frq': row.get('frq', ''),
                            'exchange': row.get('exchange', ''),
                        }

            # 保存缓存
            try:
                with open(cache_file, 'wb') as f:
                    pickle.dump(self.offline_dict, f, protocol=pickle.HIGHEST_PROTOCOL)
            except Exception:
                pass

            print(f"完成（共 {len(self.offline_dict)} 个词条）")

        except Exception as e:
            print(f"加载失败：{e}")
            self.offline_dict = {}

    # ==========================================================================
    # 词根词缀加载
    # ==========================================================================
    def _load_word_roots(self):
        if self.word_roots is not None:
            return

        root_file = None
        possible_paths = [
            'wordroot.txt',
            'ECDICT/wordroot.txt',
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'wordroot.txt'),
        ]

        for path in possible_paths:
            if os.path.exists(path):
                root_file = path
                break

        if root_file is None:
            self.word_roots = {}
            return

        try:
            with open(root_file, 'r', encoding='utf-8') as f:
                self.word_roots = json.load(f)
        except Exception:
            self.word_roots = {}

    # ==========================================================================
    # 获取单词难度等级
    # ==========================================================================
    def get_word_level(self, word):
        self._load_offline_dict()

        word_lower = word.lower()
        if word_lower not in self.offline_dict:
            return []

        tags = self.offline_dict[word_lower].get('tag', '').strip()
        if not tags:
            return []

        levels = []
        for tag in tags.split():
            if tag in self.level_labels:
                levels.append(self.level_labels[tag])

        return levels

    # ==========================================================================
    # 获取单词的词根词缀
    # ==========================================================================
    def get_word_roots(self, word):
        self._load_word_roots()

        if not self.word_roots:
            return []

        word_lower = word.lower()
        found_roots = []

        for root_key, root_info in self.word_roots.items():
            variants = [v.strip() for v in root_key.split(',')]
            for variant in variants:
                variant_clean = variant.strip('-')
                if variant_clean and variant_clean in word_lower:
                    meaning = root_info.get('meaning', '')
                    if meaning and (variant, meaning) not in found_roots:
                        found_roots.append((variant, meaning))
                    break

        return found_roots[:3]

    # ==========================================================================
    # 单词翻译
    # ==========================================================================
    def translate_word(self, word):
        word_lower = word.lower()

        if word_lower in self.local_dict:
            return self.local_dict[word_lower]

        self._load_offline_dict()
        if word_lower in self.offline_dict:
            translation = self.offline_dict[word_lower].get('translation', '')
            if translation:
                return translation.strip()

        if not self.use_online_translation:
            return "（暂无释义）"

        if requests:
            try:
                result = self._translate_youdao(word)
                if result:
                    return result
            except Exception:
                pass

            try:
                result = self._translate_iciba(word)
                if result:
                    return result
            except Exception:
                pass

        return "（暂无释义）"

    def _translate_youdao(self, word):
        if not requests:
            return None
        url = f"https://dict.youdao.com/suggest?q={word}&le=en&t=5&client=mobile"
        response = requests.get(url, headers=self.headers, timeout=5)
        if response.status_code == 200:
            data = response.json()
            if data.get('data') and data['data'].get('entries'):
                entry = data['data']['entries'][0]
                if entry.get('explain'):
                    return entry['explain']
        return None

    def _translate_iciba(self, word):
        if not requests:
            return None
        url = f"https://www.iciba.com/index.php?a=getWordMean&c=search&word={word}"
        response = requests.get(url, headers=self.headers, timeout=5)
        if response.status_code == 200:
            data = response.json()
            if data.get('baesInfo') and data['baesInfo'].get('symbols'):
                parts = data['baesInfo']['symbols'][0].get('parts', [])
                meanings = []
                for part in parts:
                    part_str = part.get('part', '')
                    means = '; '.join(part.get('means', []))
                    if part_str and means:
                        meanings.append(f"{part_str} {means}")
                    elif means:
                        meanings.append(means)
                if meanings:
                    return '\n'.join(meanings)
        return None

    # ==========================================================================
    # 文件读取
    # ==========================================================================
    def read_file(self, file_path):
        path = Path(file_path)
        suffix = path.suffix.lower()

        if not path.exists():
            raise FileNotFoundError(f"文件不存在：{file_path}")

        if suffix == '.txt':
            return self._read_txt(file_path)
        elif suffix == '.pdf':
            return self._read_pdf(file_path)
        elif suffix == '.docx':
            return self._read_docx(file_path)
        else:
            return self._read_txt(file_path)

    def _read_txt(self, file_path):
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        raise ValueError("无法识别文件编码")

    def _read_pdf(self, file_path):
        import platform
        system = platform.system()

        # 策略一：pdftotext
        try:
            import subprocess

            if system == 'Windows':
                check_cmd = 'where pdftotext'
            else:
                check_cmd = 'which pdftotext'

            result = subprocess.run(check_cmd, shell=True, capture_output=True, text=True)
            if result.returncode == 0:
                print("正在提取 PDF 文本（pdftotext）...", end=' ', flush=True)
                result = subprocess.run(
                    ['pdftotext', '-layout', file_path, '-'],
                    capture_output=True,
                    text=True
                )
                if result.returncode == 0 and result.stdout.strip():
                    print("完成")
                    return result.stdout
        except Exception:
            pass

        # 策略二：PyPDF2
        try:
            import PyPDF2
            print("正在提取 PDF 文本（PyPDF2）...", end=' ', flush=True)
            text = []
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text.append(page_text)
            result = '\n'.join(text)
            if result.strip():
                print("完成")
                return result
        except ImportError:
            pass
        except Exception:
            pass

        error_msg = """
无法读取 PDF 文件。解决方案：
1. 安装 poppler 工具包（推荐，效果好）
2. 安装 PyPDF2：pip install PyPDF2
3. 先把 PDF 转成 Word 或 txt 再分析
"""
        raise RuntimeError(error_msg)

    def _read_docx(self, file_path):
        try:
            from docx import Document
        except ImportError:
            raise ImportError("读取 Word 文档需要安装 python-docx\n运行：pip install python-docx")

        doc = Document(file_path)
        text = []
        for para in doc.paragraphs:
            if para.text.strip():
                text.append(para.text)
        return '\n'.join(text)

    # ==========================================================================
    # 进度文件相关
    # ==========================================================================
    def _get_progress_file_path(self, output_file):
        output_path = Path(output_file)
        return str(output_path.parent / f"{output_path.stem}_progress.json")

    def _save_progress(self, output_file, word_freq, translations, completed_count, input_file):
        try:
            progress_data = {
                'input_file': Path(input_file).name,
                'output_file': output_file,
                'total_words': len(word_freq),
                'completed_count': completed_count,
                'translations': dict(list(translations.items())[:completed_count]),
                'saved_at': time.strftime('%Y-%m-%d %H:%M:%S')
            }
            progress_file = self._get_progress_file_path(output_file)
            with open(progress_file, 'w', encoding='utf-8') as f:
                json.dump(progress_data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_progress(self, output_file, input_file):
        progress_file = self._get_progress_file_path(output_file)
        if not os.path.exists(progress_file):
            return None

        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if data.get('input_file') != Path(input_file).name:
                return None

            return data
        except Exception:
            return None

    def _clear_progress(self, output_file):
        try:
            progress_file = self._get_progress_file_path(output_file)
            if os.path.exists(progress_file):
                os.remove(progress_file)
        except Exception:
            pass

    # ==========================================================================
    # 导出到 Excel
    # ==========================================================================
    def export_to_excel(self, word_freq, output_path, translations, show_level=False, show_roots=False):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "词频统计"

        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=11)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ['排名', '单词', '词频', '中文释义']
        if show_level:
            headers.append('难度等级')
        if show_roots:
            headers.append('词根词缀')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for rank, (word, freq) in enumerate(word_freq.most_common(), 1):
            row = rank + 1
            translation = translations.get(word, '（暂无释义）')

            ws.cell(row=row, column=1, value=rank).alignment = center_alignment
            ws.cell(row=row, column=2, value=word).alignment = center_alignment
            ws.cell(row=row, column=3, value=freq).alignment = center_alignment
            ws.cell(row=row, column=4, value=translation).alignment = left_alignment

            col_idx = 5
            if show_level:
                levels = self.get_word_level(word)
                ws.cell(row=row, column=col_idx, value=' / '.join(levels) if levels else '').alignment = center_alignment
                col_idx += 1

            if show_roots:
                roots = self.get_word_roots(word)
                root_str = '; '.join([f"{r[0]} ({r[1]})" for r in roots])
                ws.cell(row=row, column=col_idx, value=root_str).alignment = left_alignment
                col_idx += 1

            for col in range(1, col_idx):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border

            if len(translation) > 50:
                ws.row_dimensions[row].height = 45

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50
        if show_level:
            ws.column_dimensions['E'].width = 20
        if show_roots:
            col_letter = 'F' if show_level else 'E'
            ws.column_dimensions[col_letter].width = 40

        ws.freeze_panes = 'A2'

        # === 打印优化设置 ===
        # 页面设置：A4纸，横向
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        # 页边距（厘米转英寸：1厘米 ≈ 0.3937英寸）
        ws.page_margins.left = 1.5 * 0.3937
        ws.page_margins.right = 1.5 * 0.3937
        ws.page_margins.top = 2.0 * 0.3937
        ws.page_margins.bottom = 2.0 * 0.3937
        ws.page_margins.header = 0.8 * 0.3937
        ws.page_margins.footer = 0.8 * 0.3937

        # 打印标题：每页都打印表头（第一行）
        ws.print_title_rows = '1:1'

        # 缩放：缩放到一页宽（内容多时自动缩小，保证所有列都在一页内）
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 0表示高度不限，自动分页

        # 打印网格线
        ws.print_options.gridLines = True

        wb.save(output_path)

    # ==========================================================================
    # 导出到 CSV
    # ==========================================================================
    def export_to_csv(self, word_freq, output_path, translations, show_level=False, show_roots=False):
        headers = ['排名', '单词', '词频', '中文释义']
        if show_level:
            headers.append('难度等级')
        if show_roots:
            headers.append('词根词缀')

        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for rank, (word, freq) in enumerate(word_freq.most_common(), 1):
                translation = translations.get(word, '（暂无释义）')
                row = [rank, word, freq, translation]

                if show_level:
                    levels = self.get_word_level(word)
                    row.append(' / '.join(levels) if levels else '')

                if show_roots:
                    roots = self.get_word_roots(word)
                    root_str = '; '.join([f"{r[0]} ({r[1]})" for r in roots])
                    row.append(root_str)

                writer.writerow(row)

    # ==========================================================================
    # 导出到 Anki
    # ==========================================================================
    def export_to_anki(self, word_freq, output_path, translations, max_words=None):
        words = word_freq.most_common(max_words) if max_words else word_freq.most_common()

        with open(output_path, 'w', encoding='utf-8') as f:
            for word, freq in words:
                translation = translations.get(word, '（暂无释义）')
                f.write(f"{word}\t{translation}\n")

    # ==========================================================================
    # 一键分析
    # ==========================================================================
    def analyze(self, input_file, output_file=None, enable_translation=True, max_words=None,
                resume=True, ignore_stopwords=False, lemmatize=False,
                show_level=False, show_roots=False, export_format='excel',
                filter_no_translation=False):

        print("=" * 60)
        print("英语单词词频统计工具（Pro优化版）")
        print("=" * 60)
        print()

        # 第1步：读取文件
        print(f"📖 正在读取文件：{input_file}")
        text = self.read_file(input_file)
        print(f"   读取完成，共 {len(text)} 个字符")
        print()

        # 第2步：提取单词
        print("🔍 正在提取单词...", end=' ', flush=True)
        words = self.extract_words(text)
        print(f"完成，共 {len(words)} 个单词")

        if lemmatize:
            print("🔄 正在进行词形还原...", end=' ', flush=True)
            words = [self.lemmatize_word(w) for w in words]
            print("完成")

        if ignore_stopwords:
            print("🚫 正在过滤停用词...", end=' ', flush=True)
            words = [w for w in words if w not in self.stop_words]
            print(f"完成，剩余 {len(words)} 个单词")

        print()

        # 第3步：统计词频
        print("📊 正在统计词频...", end=' ', flush=True)
        word_freq = self.count_frequency(words)
        print(f"完成，共 {len(word_freq)} 个不同单词")
        print()

        if max_words and max_words < len(word_freq):
            word_freq = Counter(dict(word_freq.most_common(max_words)))
            print(f"   （限制只处理前 {max_words} 个高频词）")
            print()

        # 第4步：生成输出文件名
        if output_file is None:
            input_path = Path(input_file)
            output_file = str(input_path.parent / f"{input_path.stem}_词频统计.xlsx")
        else:
            output_path = Path(output_file)
            if export_format == 'excel' and output_path.suffix.lower() != '.xlsx':
                output_file = str(output_path.with_suffix('.xlsx'))
            elif export_format == 'csv' and output_path.suffix.lower() != '.csv':
                output_file = str(output_path.with_suffix('.csv'))
            elif export_format == 'anki' and output_path.suffix.lower() != '.txt':
                output_file = str(output_path.with_suffix('.txt'))

        # 第5步：翻译
        translations = {}
        if enable_translation:
            self._load_offline_dict()

            need_online = 0
            for word, _ in word_freq.most_common():
                if word not in self.local_dict:
                    if self.offline_dict and word in self.offline_dict:
                        continue
                    if self.use_online_translation:
                        need_online += 1

            total_words = len(word_freq)
            print(f"📝 正在获取中文释义...")
            if self.use_online_translation:
                print(f"   共 {total_words} 个单词，其中 {need_online} 个需要在线翻译")
            else:
                print(f"   共 {total_words} 个单词（纯离线模式）")
            print()

            # 检查进度
            progress_data = None
            if resume and self.enable_progress:
                progress_data = self._load_progress(output_file, input_file)
                if progress_data:
                    print(f"📂 检测到未完成的进度：")
                    print(f"   已完成：{progress_data['completed_count']}/{progress_data['total_words']} 个单词")
                    print(f"   保存时间：{progress_data['saved_at']}")
                    print()
                    translations = progress_data.get('translations', {})

            def translate_worker(word):
                if (word not in self.local_dict and
                    (not self.offline_dict or word not in self.offline_dict) and
                    self.use_online_translation):
                    time.sleep(random.uniform(0.1, 0.3))
                return word, self.translate_word(word)

            words_to_translate = [w for w, _ in word_freq.most_common() if w not in translations]
            completed = len(translations)

            if words_to_translate:
                progress = ProgressBar(total_words, prefix="翻译进度")
                progress.update(completed)

                save_counter = 0

                with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                    futures = {executor.submit(translate_worker, word): word for word in words_to_translate}

                    for future in as_completed(futures):
                        word, translation = future.result()
                        translations[word] = translation

                        with self._progress_lock:
                            completed += 1
                            save_counter += 1
                            progress.update(completed)

                            if self.enable_progress and save_counter >= self.progress_save_interval:
                                save_counter = 0
                                self._save_progress(output_file, word_freq, translations, completed, input_file)

                print()
            else:
                print("   所有单词都已翻译完成")
                print()

            self._clear_progress(output_file)
            print(f"✅ 翻译完成，共 {len(translations)} 个单词")
            print()

        # 过滤无释义的单词
        if filter_no_translation and enable_translation:
            print("🔍 正在过滤无释义的单词...", end=' ', flush=True)
            filtered_word_freq = Counter()
            filtered_translations = {}
            for word, freq in word_freq.items():
                trans = translations.get(word, "（暂无释义）")
                if trans and trans != "（暂无释义）":
                    filtered_word_freq[word] = freq
                    filtered_translations[word] = trans
            print(f"完成，剩余 {len(filtered_word_freq)} 个单词")
            print()
            word_freq = filtered_word_freq
            translations = filtered_translations

        # 第6步：导出
        print(f"💾 正在导出文件：{output_file}")

        if export_format == 'excel':
            self.export_to_excel(word_freq, output_file, translations, show_level, show_roots)
        elif export_format == 'csv':
            self.export_to_csv(word_freq, output_file, translations, show_level, show_roots)
        elif export_format == 'anki':
            self.export_to_anki(word_freq, output_file, translations, max_words)

        print("   导出完成！")
        print()
        print("=" * 60)
        print("🎉 分析完成！")
        print(f"📊 统计单词数：{len(words)}")
        print(f"🔤 不同单词数：{len(word_freq)}")
        print(f"📁 输出文件：{output_file}")
        print("=" * 60)

        return output_file

    # ==========================================================================
    # 批量处理
    # ==========================================================================
    def batch_analyze(self, input_folder, output_folder=None, enable_translation=True,
                      max_words=None, ignore_stopwords=False, lemmatize=False,
                      show_level=False, show_roots=False, export_format='excel',
                      batch_mode='separate'):
        """批量处理文件夹中的所有文件

        Args:
            batch_mode: 'separate' - 逐个输出（每个文件单独生成结果）
                       'merge' - 合并输出（所有文件合并统计到一个总表）
        """

        input_path = Path(input_folder)
        if not input_path.exists() or not input_path.is_dir():
            raise ValueError(f"文件夹不存在：{input_folder}")

        # 支持的文件扩展名
        supported_extensions = {'.txt', '.pdf', '.docx'}

        # 扫描文件
        files = []
        for f in input_path.iterdir():
            if f.is_file() and f.suffix.lower() in supported_extensions:
                files.append(f)

        if not files:
            print(f"❌ 在文件夹中没有找到支持的文件（支持：.txt, .pdf, .docx）")
            return []

        print(f"📁 找到 {len(files)} 个文件待处理")
        print(f"📊 批量模式：{'逐个输出' if batch_mode == 'separate' else '合并输出'}")
        print()

        # 输出文件夹
        if output_folder is None:
            output_folder = str(input_path / "词频统计结果")
        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)

        # === 合并输出模式 ===
        if batch_mode == 'merge':
            print("=" * 60)
            print("🔗 合并模式：将所有文件合并统计")
            print("=" * 60)
            print()

            all_words = []

            for i, file_path in enumerate(files, 1):
                print(f"📄 正在读取第 {i}/{len(files)} 个文件：{file_path.name}")
                try:
                    text = self.read_file(str(file_path))
                    words = self.extract_words(text)

                    if lemmatize:
                        words = [self.lemmatize_word(w) for w in words]

                    if ignore_stopwords:
                        words = [w for w in words if w not in self.stop_words]

                    all_words.extend(words)
                    print(f"   完成，{len(words)} 个单词")
                except Exception as e:
                    print(f"   ❌ 读取失败：{e}")

            print()
            print(f"📊 合并完成，共 {len(all_words)} 个单词")
            print()

            # 统计词频
            print("📊 正在统计词频...", end=' ', flush=True)
            word_freq = self.count_frequency(all_words)
            print(f"完成，共 {len(word_freq)} 个不同单词")
            print()

            if max_words and max_words < len(word_freq):
                word_freq = Counter(dict(word_freq.most_common(max_words)))
                print(f"   （限制只处理前 {max_words} 个高频词）")
                print()

            # 生成输出文件名
            if export_format == 'excel':
                output_file = str(output_path / "合并词频统计.xlsx")
            elif export_format == 'csv':
                output_file = str(output_path / "合并词频统计.csv")
            else:
                output_file = str(output_path / "合并生词本.txt")

            # 翻译
            translations = {}
            if enable_translation:
                self._load_offline_dict()

                need_online = 0
                for word, _ in word_freq.most_common():
                    if word not in self.local_dict:
                        if self.offline_dict and word in self.offline_dict:
                            continue
                        if self.use_online_translation:
                            need_online += 1

                total_words = len(word_freq)
                print(f"📝 正在获取中文释义...")
                if self.use_online_translation:
                    print(f"   共 {total_words} 个单词，其中 {need_online} 个需要在线翻译")
                else:
                    print(f"   共 {total_words} 个单词（纯离线模式）")
                print()

                def translate_worker(word):
                    if (word not in self.local_dict and
                        (not self.offline_dict or word not in self.offline_dict) and
                        self.use_online_translation):
                        time.sleep(random.uniform(0.1, 0.3))
                    return word, self.translate_word(word)

                words_to_translate = [w for w, _ in word_freq.most_common()]
                completed = 0

                if words_to_translate:
                    progress = ProgressBar(total_words, prefix="翻译进度")

                    with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                        futures = {executor.submit(translate_worker, word): word for word in words_to_translate}

                        for future in as_completed(futures):
                            word, translation = future.result()
                            translations[word] = translation

                            with self._progress_lock:
                                completed += 1
                                progress.update(completed)

                    print()

                print(f"✅ 翻译完成，共 {len(translations)} 个单词")
                print()

            # 导出
            print(f"💾 正在导出文件：{output_file}")

            if export_format == 'excel':
                self.export_to_excel(word_freq, output_file, translations, show_level, show_roots)
            elif export_format == 'csv':
                self.export_to_csv(word_freq, output_file, translations, show_level, show_roots)
            elif export_format == 'anki':
                self.export_to_anki(word_freq, output_file, translations, max_words)

            print("   导出完成！")
            print()
            print("=" * 60)
            print("🎉 批量合并处理完成！")
            print(f"📊 总单词数：{len(all_words)}")
            print(f"🔤 不同单词数：{len(word_freq)}")
            print(f"📁 输出文件：{output_file}")
            print("=" * 60)

            return [output_file]

        # === 逐个输出模式 ===
        else:
            results = []
            success_count = 0
            fail_count = 0

            for i, file_path in enumerate(files, 1):
                print(f"{'=' * 60}")
                print(f"📄 正在处理第 {i}/{len(files)} 个文件：{file_path.name}")
                print(f"{'=' * 60}")

                try:
                    # 生成输出文件名
                    if export_format == 'excel':
                        output_file = str(output_path / f"{file_path.stem}_词频统计.xlsx")
                    elif export_format == 'csv':
                        output_file = str(output_path / f"{file_path.stem}_词频统计.csv")
                    else:
                        output_file = str(output_path / f"{file_path.stem}_生词本.txt")

                    # 调用单个文件分析
                    result = self.analyze(
                        input_file=str(file_path),
                        output_file=output_file,
                        enable_translation=enable_translation,
                        max_words=max_words,
                        resume=False,
                        ignore_stopwords=ignore_stopwords,
                        lemmatize=lemmatize,
                        show_level=show_level,
                        show_roots=show_roots,
                        export_format=export_format,
                    )
                    results.append(result)
                    success_count += 1
                    print()

                except Exception as e:
                    print(f"❌ 处理失败：{e}")
                    fail_count += 1
                    print()

            # 汇总
            print("=" * 60)
            print("🎉 批量处理完成！")
            print(f"   成功：{success_count} 个")
            print(f"   失败：{fail_count} 个")
            print(f"   输出文件夹：{output_folder}")
            print("=" * 60)

            return results


# ==============================================================================
# 主函数
# ==============================================================================
def main():
    config = ConfigManager()

    print("=" * 60)
    print("英语单词词频统计工具（Pro优化版）")
    print("=" * 60)
    print()

    try:
        # 1. 输入文件
        default_input = config.get('last_input_file', '')
        if default_input:
            input_file = input(f"请输入要分析的文件路径（回车使用上次：{default_input}）：").strip()
            if not input_file:
                input_file = default_input
        else:
            input_file = input("请输入要分析的文件路径：").strip()

        if not input_file:
            print("错误：文件路径不能为空")
            return

        input_file = input_file.strip('"').strip("'")

        if not os.path.exists(input_file):
            print(f"错误：文件不存在 - {input_file}")
            return

        config.set('last_input_file', input_file)

        # 2. 输出文件
        default_output = config.get('last_output_path', '')
        if default_output:
            output_file = input(f"请输入保存路径（回车使用默认/上次）：").strip()
        else:
            output_file = input("请输入保存路径（回车使用默认）：").strip()

        if output_file:
            output_file = output_file.strip('"').strip("'")
            config.set('last_output_path', output_file)
        else:
            output_file = None

        # 3. 导出格式
        default_format = config.get('export_format', 'excel')
        print()
        print("请选择导出格式：")
        print("  1. Excel（.xlsx）- 默认")
        print("  2. CSV（.csv）")
        print("  3. Anki 生词本（.txt）")
        format_choice = input(f"请选择（1/2/3，回车默认 {default_format}）：").strip()

        if format_choice == '2':
            export_format = 'csv'
        elif format_choice == '3':
            export_format = 'anki'
        else:
            export_format = default_format

        config.set('export_format', export_format)

        # 4. 是否翻译
        print()
        enable_translation = input("是否获取中文释义？(Y/n，默认是)：").strip().lower()
        enable_translation = enable_translation != 'n'

        # 5. 纯离线模式
        if enable_translation:
            print()
            use_offline = input("是否使用纯离线模式？(y/N，默认否)：").strip().lower()
            use_offline = use_offline == 'y'
        else:
            use_offline = False

        # 6. 忽略停用词
        print()
        default_stop = config.get('ignore_stopwords', False)
        stop_str = '是' if default_stop else '否'
        ignore_stop = input(f"是否忽略停用词？(y/N，上次：{stop_str})：").strip().lower()
        ignore_stop = ignore_stop == 'y'
        config.set('ignore_stopwords', ignore_stop)

        # 7. 词形还原
        print()
        default_lemma = config.get('lemmatize', False)
        lemma_str = '是' if default_lemma else '否'
        lemmatize = input(f"是否进行词形还原？(y/N，上次：{lemma_str})：").strip().lower()
        lemmatize = lemmatize == 'y'
        config.set('lemmatize', lemmatize)

        # 8. 难度等级
        print()
        show_level = input("是否显示单词难度等级（四级/六级/考研等）？(y/N，默认否)：").strip().lower()
        show_level = show_level == 'y'

        # 9. 词根词缀
        print()
        show_roots = input("是否显示词根词缀分析？(y/N，默认否)：").strip().lower()
        show_roots = show_roots == 'y'

        # 10. 线程数
        print()
        default_workers = config.get('max_workers', 3)
        workers_str = input(f"使用多少个线程？（回车默认 {default_workers} 个）：").strip()
        try:
            max_workers = int(workers_str) if workers_str else default_workers
        except ValueError:
            max_workers = default_workers
        config.set('max_workers', max_workers)

        # 11. 最多单词数
        print()
        max_words_str = input("最多翻译多少个单词？（回车默认全部）：").strip()
        max_words = int(max_words_str) if max_words_str else None

        print()

        config.save()

        # 创建分析器
        analyzer = WordFrequencyAnalyzer()
        analyzer.use_online_translation = not use_offline
        analyzer.max_workers = max_workers

        # 检查是否继续进度
        resume = True
        if output_file:
            progress_file = analyzer._get_progress_file_path(output_file)
            if os.path.exists(progress_file):
                print()
                resume_choice = input("检测到未完成的进度，是否继续？(Y/n，默认是)：").strip().lower()
                resume = resume_choice != 'n'

        # 执行分析
        analyzer.analyze(
            input_file=input_file,
            output_file=output_file,
            enable_translation=enable_translation,
            max_words=max_words,
            resume=resume,
            ignore_stopwords=ignore_stop,
            lemmatize=lemmatize,
            show_level=show_level,
            show_roots=show_roots,
            export_format=export_format,
            filter_no_translation=filter_no_trans,
        )

    except KeyboardInterrupt:
        print("\n\n用户中断，程序退出。")
    except Exception as e:
        print(f"\n❌ 错误：{type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print()
        try:
            input("按回车键退出...")
        except EOFError:
            pass


# ==============================================================================
# GUI 界面类
# ==============================================================================
class WordFrequencyGUI:
    """词频统计工具的图形界面"""

    def __init__(self, root):
        self.root = root
        
        # 语言管理器
        self.lang = LanguageManager('zh')
        
        self.root.title(self.lang.get('app_title'))
        self.root.geometry("700x680")
        self.root.resizable(True, True)

        # 配置管理器
        self.config = ConfigManager()

        # 分析器
        self.analyzer = None
        self.is_running = False

        # 创建界面
        self._create_widgets()

        # 加载上次的配置
        self._load_config()

    def _create_widgets(self):
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # === 文件选择区域 ===
        file_frame = ttk.LabelFrame(main_frame, text=self.lang.get("file_settings"), padding="10")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)

        # 批量模式切换
        self.batch_mode_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(file_frame, text=self.lang.get("batch_mode"),
                        variable=self.batch_mode_var,
                        command=self._toggle_batch_mode).grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))

        # 批量输出模式选择（默认隐藏，勾选批量模式后显示）
        self.batch_output_frame = ttk.Frame(file_frame)
        self.batch_output_mode_var = tk.StringVar(value="separate")
        ttk.Label(self.batch_output_frame, text=self.lang.get("output_mode")).grid(row=0, column=0, padx=(0, 10))
        ttk.Radiobutton(self.batch_output_frame, text=self.lang.get("separate_output"),
                        variable=self.batch_output_mode_var, value="separate").grid(row=0, column=1, padx=5)
        ttk.Radiobutton(self.batch_output_frame, text=self.lang.get("merge_output"),
                        variable=self.batch_output_mode_var, value="merge").grid(row=0, column=2, padx=5)

        # 输入文件/文件夹
        self.input_label = ttk.Label(file_frame, text=self.lang.get("input_file"))
        self.input_label.grid(row=2, column=0, sticky=tk.W, pady=5)
        self.input_var = tk.StringVar()
        self.input_entry = ttk.Entry(file_frame, textvariable=self.input_var)
        self.input_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        self.input_browse_btn = ttk.Button(file_frame, text=self.lang.get("browse"), command=self._browse_input)
        self.input_browse_btn.grid(row=2, column=2, pady=5)

        # 输出文件/文件夹
        self.output_label = ttk.Label(file_frame, text=self.lang.get("output_file"))
        self.output_label.grid(row=3, column=0, sticky=tk.W, pady=5)
        self.output_var = tk.StringVar()
        self.output_entry = ttk.Entry(file_frame, textvariable=self.output_var)
        self.output_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        self.output_browse_btn = ttk.Button(file_frame, text=self.lang.get("browse"), command=self._browse_output)
        self.output_browse_btn.grid(row=3, column=2, pady=5)

        # 词典文件
        ttk.Label(file_frame, text=self.lang.get("dict_file")).grid(row=4, column=0, sticky=tk.W, pady=5)
        self.dict_path_var = tk.StringVar()
        self.dict_entry = ttk.Entry(file_frame, textvariable=self.dict_path_var)
        self.dict_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Button(file_frame, text=self.lang.get("browse"), command=self._browse_dict).grid(row=4, column=2, pady=5)

        # === 选项区域 ===
        options_frame = ttk.LabelFrame(main_frame, text=self.lang.get("analysis_options"), padding="10")
        options_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        options_frame.columnconfigure(0, weight=1)
        options_frame.columnconfigure(1, weight=1)

        # 第一列
        self.ignore_stop_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="忽略停用词", variable=self.ignore_stop_var).grid(row=0, column=0, sticky=tk.W, pady=3)

        self.lemmatize_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="词形还原", variable=self.lemmatize_var).grid(row=1, column=0, sticky=tk.W, pady=3)

        self.offline_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="纯离线模式（不联网）", variable=self.offline_var).grid(row=2, column=0, sticky=tk.W, pady=3)

        self.enable_trans_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="获取中文释义", variable=self.enable_trans_var).grid(row=3, column=0, sticky=tk.W, pady=3)

        self.filter_no_trans_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="只保留有释义的单词", variable=self.filter_no_trans_var).grid(row=4, column=0, sticky=tk.W, pady=3)

        # 第二列
        self.show_level_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="显示难度等级", variable=self.show_level_var).grid(row=0, column=1, sticky=tk.W, pady=3)

        self.show_roots_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="显示词根词缀", variable=self.show_roots_var).grid(row=1, column=1, sticky=tk.W, pady=3)

        # 线程数
        ttk.Label(options_frame, text="线程数：").grid(row=2, column=1, sticky=tk.W, pady=3)
        self.workers_var = tk.StringVar(value="3")
        workers_spin = ttk.Spinbox(options_frame, from_=1, to=10, width=5, textvariable=self.workers_var)
        workers_spin.grid(row=2, column=1, sticky=tk.E, pady=3)

        # 最多单词数
        ttk.Label(options_frame, text="最多翻译：").grid(row=3, column=1, sticky=tk.W, pady=3)
        self.max_words_var = tk.StringVar(value="")
        max_words_entry = ttk.Entry(options_frame, width=10, textvariable=self.max_words_var)
        max_words_entry.grid(row=3, column=1, sticky=tk.E, pady=3)

        # === 导出格式 ===
        format_frame = ttk.LabelFrame(main_frame, text=self.lang.get("export_format"), padding="10")
        format_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        self.format_var = tk.StringVar(value="excel")
        ttk.Radiobutton(format_frame, text="Excel (.xlsx)", variable=self.format_var, value="excel").grid(row=0, column=0, padx=10)
        ttk.Radiobutton(format_frame, text="CSV (.csv)", variable=self.format_var, value="csv").grid(row=0, column=1, padx=10)
        ttk.Radiobutton(format_frame, text="Anki 生词本 (.txt)", variable=self.format_var, value="anki").grid(row=0, column=2, padx=10)

        # === 按钮区域 ===
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=3, column=0, pady=(0, 10))

        self.start_button = ttk.Button(btn_frame, text=self.lang.get('start_analysis'), command=self._start_analysis)
        self.start_button.grid(row=0, column=0, padx=5)

        self.about_button = ttk.Button(btn_frame, text=self.lang.get('about_btn'), command=self._show_about)
        self.about_button.grid(row=0, column=1, padx=5)

        self.lang_button = ttk.Button(btn_frame, text=self.lang.get('language_btn'), command=self._toggle_language)
        self.lang_button.grid(row=0, column=2, padx=5)

        # === 进度条 ===
        progress_frame = ttk.LabelFrame(main_frame, text=self.lang.get("progress"), padding="10")
        progress_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)

        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)

        self.status_var = tk.StringVar(value="准备就绪")
        ttk.Label(progress_frame, textvariable=self.status_var).grid(row=1, column=0, sticky=tk.W)

        # === 日志输出 ===
        log_frame = ttk.LabelFrame(main_frame, text=self.lang.get("run_log"), padding="10")
        log_frame.grid(row=5, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.log_text.config(state=tk.DISABLED)

    def _load_config(self):
        """加载配置到界面"""
        # 加载语言设置
        saved_lang = self.config.get('language', 'zh')
        if saved_lang in ['zh', 'en']:
            self.lang.set_language(saved_lang)
            self.root.title(self.lang.get('app_title'))
            self.start_button.config(text=self.lang.get('start_analysis'))
            self.about_button.config(text=self.lang.get('about_btn'))
            self.lang_button.config(text=self.lang.get('language_btn'))
        
        self.ignore_stop_var.set(self.config.get('ignore_stopwords', False))
        self.lemmatize_var.set(self.config.get('lemmatize', False))
        self.workers_var.set(str(self.config.get('max_workers', 3)))
        self.format_var.set(self.config.get('export_format', 'excel'))

        last_input = self.config.get('last_input_file', '')
        if last_input:
            self.input_var.set(last_input)

        last_output = self.config.get('last_output_path', '')
        if last_output:
            self.output_var.set(last_output)

        last_dict = self.config.get('dict_path', '')
        if last_dict:
            self.dict_path_var.set(last_dict)

        self.filter_no_trans_var.set(self.config.get('filter_no_translation', False))

    def _save_config(self):
        """保存界面配置"""
        self.config.set('ignore_stopwords', self.ignore_stop_var.get())
        self.config.set('lemmatize', self.lemmatize_var.get())
        self.config.set('max_workers', int(self.workers_var.get()))
        self.config.set('export_format', self.format_var.get())
        self.config.set('last_input_file', self.input_var.get())
        self.config.set('last_output_path', self.output_var.get())
        self.config.set('dict_path', self.dict_path_var.get())
        self.config.set('filter_no_translation', self.filter_no_trans_var.get())
        self.config.save()

    def _toggle_batch_mode(self):
        """切换批量处理模式"""
        if self.batch_mode_var.get():
            self.input_label.config(text="输入文件夹：")
            self.output_label.config(text="输出文件夹：")
            # 显示输出模式选择
            self.batch_output_frame.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(0, 5))
        else:
            self.input_label.config(text="输入文件：")
            self.output_label.config(text="输出文件：")
            # 隐藏输出模式选择
            self.batch_output_frame.grid_remove()

    def _browse_input(self):
        if self.batch_mode_var.get():
            # 批量模式：选择文件夹
            folder = filedialog.askdirectory(title="选择包含文件的文件夹")
            if folder:
                self.input_var.set(folder)
        else:
            # 单文件模式：选择文件
            filename = filedialog.askopenfilename(
                title="选择要分析的文件",
                filetypes=[
                    ("所有支持的文件", "*.txt *.pdf *.docx"),
                    ("文本文件", "*.txt"),
                    ("PDF 文件", "*.pdf"),
                    ("Word 文档", "*.docx"),
                    ("所有文件", "*.*")
                ]
            )
            if filename:
                self.input_var.set(filename)

    def _browse_output(self):
        if self.batch_mode_var.get():
            # 批量模式：选择输出文件夹
            folder = filedialog.askdirectory(title="选择输出文件夹")
            if folder:
                self.output_var.set(folder)
        else:
            # 单文件模式：选择保存文件
            export_format = self.format_var.get()
            if export_format == 'excel':
                defaultext = '.xlsx'
                filetypes = [("Excel 文件", "*.xlsx")]
            elif export_format == 'csv':
                defaultext = '.csv'
                filetypes = [("CSV 文件", "*.csv")]
            else:
                defaultext = '.txt'
                filetypes = [("文本文件", "*.txt")]

            filename = filedialog.asksaveasfilename(
                title="保存文件",
                defaultextension=defaultext,
                filetypes=filetypes
            )
            if filename:
                self.output_var.set(filename)

    def _browse_dict(self):
        """浏览选择词典文件"""
        filename = filedialog.askopenfilename(
            title="选择词典文件",
            filetypes=[
                ("CSV 词典文件", "*.csv"),
                ("所有文件", "*.*")
            ]
        )
        if filename:
            self.dict_path_var.set(filename)

    def _log(self, message):
        """添加日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def _update_progress(self, value, status=""):
        """更新进度条"""
        self.progress_var.set(value)
        if status:
            self.status_var.set(status)
        self.root.update_idletasks()

    def _toggle_language(self):
        """切换语言"""
        if self.lang.language == 'zh':
            self.lang.set_language('en')
        else:
            self.lang.set_language('zh')
        self._update_ui_language()
        self.config.set('language', self.lang.language)
        self.config.save()
    
    def _update_ui_language(self):
        """更新界面语言"""
        self.root.title(self.lang.get('app_title'))
        self.start_button.config(text=self.lang.get('start_analysis'))
        self.about_button.config(text=self.lang.get('about_btn'))
        self.lang_button.config(text=self.lang.get('language_btn'))
        
        # 更新标签
        if self.batch_mode_var.get():
            self.input_label.config(text=self.lang.get('input_folder'))
            self.output_label.config(text=self.lang.get('output_folder'))
        else:
            self.input_label.config(text=self.lang.get('input_file'))
            self.output_label.config(text=self.lang.get('output_file'))
        
        # 更新所有复选框和标签
        self._update_labels()
    
    def _update_labels(self):
        """更新所有标签文字（简化版）"""
        pass

    def _show_about(self):
        """显示公告/关于对话框"""
        messagebox.showinfo(self.lang.get('about_title'), self.lang.get('about_content'))

    def _start_analysis(self):
        if self.is_running:
            return

        input_path = self.input_var.get().strip()
        if not input_path:
            if self.batch_mode_var.get():
                messagebox.showerror("错误", "请选择输入文件夹")
            else:
                messagebox.showerror("错误", "请选择输入文件")
            return

        if not os.path.exists(input_path):
            if self.batch_mode_var.get():
                messagebox.showerror("错误", f"文件夹不存在：{input_path}")
            else:
                messagebox.showerror("错误", f"文件不存在：{input_path}")
            return

        if self.batch_mode_var.get() and not os.path.isdir(input_path):
            messagebox.showerror("错误", "批量模式下请选择文件夹")
            return

        # 保存配置
        self._save_config()

        # 禁用按钮
        self.is_running = True
        self.start_button.config(state=tk.DISABLED)
        self.progress_var.set(0)
        self.status_var.set("正在准备...")
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

        # 在后台线程运行
        thread = threading.Thread(target=self._run_analysis, daemon=True)
        thread.start()

    def _run_analysis(self):
        try:
            # 创建分析器
            dict_path = self.dict_path_var.get().strip() or None
            self.analyzer = WordFrequencyAnalyzer(dict_path=dict_path)
            self.analyzer.use_online_translation = not self.offline_var.get()
            self.analyzer.max_workers = int(self.workers_var.get())

            # 获取参数
            input_path = self.input_var.get().strip()
            output_path = self.output_var.get().strip() or None
            enable_trans = self.enable_trans_var.get()
            ignore_stop = self.ignore_stop_var.get()
            lemmatize = self.lemmatize_var.get()
            show_level = self.show_level_var.get()
            show_roots = self.show_roots_var.get()
            export_format = self.format_var.get()
            filter_no_trans = self.filter_no_trans_var.get()
            is_batch = self.batch_mode_var.get()

            max_words_str = self.max_words_var.get().strip()
            max_words = int(max_words_str) if max_words_str else None

            # 重定向输出到日志
            import io
            old_stdout = sys.stdout
            sys.stdout = LogRedirector(self._log)

            try:
                if is_batch:
                    # 批量处理模式
                    result = self.analyzer.batch_analyze(
                        input_folder=input_path,
                        output_folder=output_path,
                        enable_translation=enable_trans,
                        max_words=max_words,
                        ignore_stopwords=ignore_stop,
                        lemmatize=lemmatize,
                        show_level=show_level,
                        show_roots=show_roots,
                        export_format=export_format,
                        batch_mode=self.batch_output_mode_var.get(),
                    )

                    # 更新进度到100%
                    self._update_progress(100, "完成！")

                    if result:
                        self.root.after(0, lambda: messagebox.showinfo(
                            "完成",
                            f"批量处理完成！\n成功处理 {len(result)} 个文件\n输出文件夹：{output_path or input_path + '/词频统计结果'}"
                        ))
                    else:
                        self.root.after(0, lambda: messagebox.showwarning("提示", "没有找到可处理的文件"))
                else:
                    # 单文件模式
                    result = self.analyzer.analyze(
                        input_file=input_path,
                        output_file=output_path,
                        enable_translation=enable_trans,
                        max_words=max_words,
                        resume=True,
                        ignore_stopwords=ignore_stop,
                        lemmatize=lemmatize,
                        show_level=show_level,
                        show_roots=show_roots,
                        export_format=export_format,
                        filter_no_translation=filter_no_trans,
                    )

                    # 更新进度到100%
                    self._update_progress(100, "完成！")

                    self.root.after(0, lambda: messagebox.showinfo("完成", f"分析完成！\n输出文件：{result}"))

            finally:
                sys.stdout = old_stdout

        except Exception as e:
            error_msg = f"错误：{type(e).__name__}: {e}"
            self._log(error_msg)
            self.root.after(0, lambda: messagebox.showerror("错误", str(e)))

        finally:
            self.is_running = False
            self.root.after(0, lambda: self.start_button.config(state=tk.NORMAL))


class LogRedirector:
    """将print输出重定向到GUI日志"""

    def __init__(self, log_func):
        self.log_func = log_func
        self.buffer = ""

    def write(self, text):
        self.buffer += text
        if "\n" in self.buffer:
            lines = self.buffer.split("\n")
            for line in lines[:-1]:
                if line.strip():
                    self.log_func(line)
            self.buffer = lines[-1]

    def flush(self):
        if self.buffer.strip():
            self.log_func(self.buffer)
            self.buffer = ""


# ==============================================================================
# 主入口
# ==============================================================================
def main():
    # 如果有命令行参数，使用命令行模式
    if len(sys.argv) > 1:
        main_cli()
        return

    # 尝试启动 GUI
    if GUI_AVAILABLE:
        try:
            root = tk.Tk()
            app = WordFrequencyGUI(root)
            root.mainloop()
            return
        except Exception as e:
            print(f"GUI 启动失败，使用命令行模式：{e}")

    # 降级到命令行模式
    main_cli()


def main_cli():
    """命令行模式主函数"""
    config = ConfigManager()

    print("=" * 60)
    print("英语单词词频统计工具（Pro优化版）")
    print("=" * 60)
    print()

    try:
        # 1. 输入文件
        default_input = config.get('last_input_file', '')
        if default_input:
            input_file = input(f"请输入要分析的文件路径（回车使用上次：{default_input}）：").strip()
            if not input_file:
                input_file = default_input
        else:
            input_file = input("请输入要分析的文件路径：").strip()

        if not input_file:
            print("错误：文件路径不能为空")
            return

        input_file = input_file.strip('"').strip("'")

        if not os.path.exists(input_file):
            print(f"错误：文件不存在 - {input_file}")
            return

        config.set('last_input_file', input_file)

        # 2. 输出文件
        default_output = config.get('last_output_path', '')
        if default_output:
            output_file = input(f"请输入保存路径（回车使用默认/上次）：").strip()
        else:
            output_file = input("请输入保存路径（回车使用默认）：").strip()

        if output_file:
            output_file = output_file.strip('"').strip("'")
            config.set('last_output_path', output_file)
        else:
            output_file = None

        # 3. 导出格式
        default_format = config.get('export_format', 'excel')
        print()
        print("请选择导出格式：")
        print("  1. Excel（.xlsx）- 默认")
        print("  2. CSV（.csv）")
        print("  3. Anki 生词本（.txt）")
        format_choice = input(f"请选择（1/2/3，回车默认 {default_format}）：").strip()

        if format_choice == '2':
            export_format = 'csv'
        elif format_choice == '3':
            export_format = 'anki'
        else:
            export_format = default_format

        config.set('export_format', export_format)

        # 4. 是否翻译
        print()
        enable_translation = input("是否获取中文释义？(Y/n，默认是)：").strip().lower()
        enable_translation = enable_translation != 'n'

        # 5. 纯离线模式
        if enable_translation:
            print()
            use_offline = input("是否使用纯离线模式？(y/N，默认否)：").strip().lower()
            use_offline = use_offline == 'y'
        else:
            use_offline = False

        # 6. 忽略停用词
        print()
        default_stop = config.get('ignore_stopwords', False)
        stop_str = '是' if default_stop else '否'
        ignore_stop = input(f"是否忽略停用词？(y/N，上次：{stop_str})：").strip().lower()
        ignore_stop = ignore_stop == 'y'
        config.set('ignore_stopwords', ignore_stop)

        # 7. 词形还原
        print()
        default_lemma = config.get('lemmatize', False)
        lemma_str = '是' if default_lemma else '否'
        lemmatize = input(f"是否进行词形还原？(y/N，上次：{lemma_str})：").strip().lower()
        lemmatize = lemmatize == 'y'
        config.set('lemmatize', lemmatize)

        # 8. 难度等级
        print()
        show_level = input("是否显示单词难度等级？(Y/n，默认是)：").strip().lower()
        show_level = show_level != 'n'

        # 9. 词根词缀
        print()
        show_roots = input("是否显示词根词缀分析？(Y/n，默认是)：").strip().lower()
        show_roots = show_roots != 'n'

        # 9.5 只保留有释义的单词
        print()
        filter_no_trans = input("是否只保留有释义的单词？(y/N，默认否)：").strip().lower()
        filter_no_trans = filter_no_trans == 'y'

        # 10. 线程数
        print()
        default_workers = config.get('max_workers', 3)
        workers_str = input(f"使用多少个线程？（回车默认 {default_workers} 个）：").strip()
        try:
            max_workers = int(workers_str) if workers_str else default_workers
        except ValueError:
            max_workers = default_workers
        config.set('max_workers', max_workers)

        # 11. 最多单词数
        print()
        max_words_str = input("最多翻译多少个单词？（回车默认全部）：").strip()
        max_words = int(max_words_str) if max_words_str else None

        print()

        config.save()

        # 创建分析器
        analyzer = WordFrequencyAnalyzer()
        analyzer.use_online_translation = not use_offline
        analyzer.max_workers = max_workers

        # 检查是否继续进度
        resume = True
        if output_file:
            progress_file = analyzer._get_progress_file_path(output_file)
            if os.path.exists(progress_file):
                print()
                resume_choice = input("检测到未完成的进度，是否继续？(Y/n，默认是)：").strip().lower()
                resume = resume_choice != 'n'

        # 执行分析
        analyzer.analyze(
            input_file=input_file,
            output_file=output_file,
            enable_translation=enable_translation,
            max_words=max_words,
            resume=resume,
            ignore_stopwords=ignore_stop,
            lemmatize=lemmatize,
            show_level=show_level,
            show_roots=show_roots,
            export_format=export_format,
            filter_no_translation=filter_no_trans,
        )

    except KeyboardInterrupt:
        print("\n\n用户中断，程序退出。")
    except Exception as e:
        print(f"\n❌ 错误：{type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print()
        try:
            input("按回车键退出...")
        except EOFError:
            pass


if __name__ == '__main__':
    main()
